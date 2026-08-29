"""首页四档证件去向 + 盘库清单：单位统一成「本」，数字能拿去和柜子核对。

首页这一行原来数的是**出国申请条数**，标签却写着「证件在库 / 领用中 / 逾期未还」，
于是三样都不对：

- 没提过出国申请的人，他的证在保管处躺着，一本都没被数进去；
- 一个人持三本、借走一本，按申请算就整个人算「领用中」，另外两本凭空消失；
- 路径B（做证自办）那条申请卡在「在库」桶里，可证明明在本人手上——结果
  **「逾期」比「领用中」还大**，逾期本该是它的子集。

现在四档全部按「本」，靠两个恒等式撑着：
    在库 + 借出未还 = 在控人员台账登记的总本数
    逾期 ⊆ 借出未还 + 新办未入库
「新办未入库」必须单独一档：那本证还没进台账、也没进过柜子，塞进前两档任何一档
都会把对账恒等式打破。
"""
import re
import sqlite3
from datetime import datetime, timedelta

import pytest

from config import Config

_CSRF = re.compile(r'name="csrf-token" content="([^"]+)"')
_VALID_ID = "110101199001012133"


@pytest.fixture()
def c(tmp_path, monkeypatch):
    """四个人，覆盖盘库会遇到的每一种情况：

    甲 在控，持 3 本（护照/港澳/台湾），护照凭领用单借出未还且已逾期
    乙 在控，持 1 本，没借过 —— 证一直在柜子里
    丙 已撤控，台账还留着一行 —— 证已随撤控移交，不在柜子里
    丁 在控，路径B 做证自办，行程早已结束、新证还没交回 —— 在外且已逾期
    戊 在控，路径B 做证自办，下个月才出发 —— 在外但还没到期
    """
    monkeypatch.setattr(Config, "DATABASE", str(tmp_path / "t.db"))
    up = tmp_path / "up"; up.mkdir()
    monkeypatch.setattr(Config, "UPLOAD_FOLDER", str(up))
    monkeypatch.setattr(Config, "EXPORT_FOLDER", str(tmp_path / "exp"))
    monkeypatch.setattr(Config, "BACKUP_FOLDER", str(tmp_path / "bak"))
    import database
    database.init_db(); database.run_migrations(); database.seed_data()

    ago = (datetime.now() - timedelta(days=120)).strftime("%Y%m%d")
    soon = (datetime.now() + timedelta(days=30)).strftime("%Y%m%d")
    db = sqlite3.connect(Config.DATABASE)

    def person(pid, nm, status="active"):
        db.execute("INSERT INTO personnel_filing (id,surname,given_name,gender,birth_date,"
                   "id_number,residence,political_status,work_unit,position_or_title,"
                   "supervisor_unit,status,operator) VALUES (?,?,'','男','19900101',?,"
                   "'浙江宁波市鄞州区','群众','总部','科长','人事处',?,'admin')",
                   (pid, nm, _VALID_ID, status))

    person(1, "甲")
    db.execute("INSERT INTO certificates (id,personnel_filing_id,unit,department,name,"
               "passport_no,passport_expiry,passport_submit_date,"
               "hm_pass_no,hm_pass_expiry,hm_pass_submit_date,"
               "tw_pass_no,tw_pass_expiry,tw_pass_submit_date,operator) VALUES "
               "(1,1,'总部','技术部','甲','E1','20351231','20250101',"
               "'C1','20351231','20250101','T1','20351231','20250101','admin')")
    db.execute("INSERT INTO travel_details (id,personnel_filing_id,unit,department,name,position,"
               "id_number,destination_passport,category,travel_dates,travel_start,travel_end,"
               "need_new_passport,passport_collect_date,passport_no,operator) VALUES "
               "(1,1,'总部','技术部','甲','科长',?,'美国/护照','01','历史',?,?,'否',?,'E1','admin')",
               (_VALID_ID, ago, ago, ago))
    db.execute("INSERT INTO cert_issuance (id,travel_id,personnel_filing_id,holder_name,id_number,"
               "cert_types,cert_nos,issue_date,issuer,status,operator) VALUES "
               "(1,1,1,'甲',?,'01','E1',?,'保管处','issued','admin')", (_VALID_ID, ago))

    person(2, "乙")
    db.execute("INSERT INTO certificates (id,personnel_filing_id,unit,department,name,"
               "passport_no,passport_expiry,passport_submit_date,operator) "
               "VALUES (2,2,'总部','技术部','乙','E2','20351231','20250101','admin')")

    person(3, "丙", "decontrolled")
    db.execute("INSERT INTO certificates (id,personnel_filing_id,unit,department,name,"
               "passport_no,passport_expiry,passport_submit_date,operator) "
               "VALUES (3,3,'总部','技术部','丙','E3','20351231','20250101','admin')")

    person(4, "丁")
    db.execute("INSERT INTO travel_details (id,personnel_filing_id,unit,department,name,position,"
               "id_number,destination_passport,category,travel_dates,travel_start,travel_end,"
               "need_new_passport,operator) VALUES "
               "(4,4,'总部','技术部','丁','科长',?,'美国/护照','01','历史',?,?,'是','admin')",
               (_VALID_ID, ago, ago))

    person(5, "戊")
    db.execute("INSERT INTO travel_details (id,personnel_filing_id,unit,department,name,position,"
               "id_number,destination_passport,category,travel_dates,travel_start,travel_end,"
               "need_new_passport,operator) VALUES "
               "(5,5,'总部','技术部','戊','科长',?,'美国/护照','01','下月',?,?,'是','admin')",
               (_VALID_ID, soon, soon))
    db.commit(); db.close()

    from app import create_app
    cl = create_app().test_client()
    tok = _CSRF.search(cl.get("/login").get_data(as_text=True)).group(1)
    cl.post("/login", data={"username": "admin", "password": "admin123", "csrf_token": tok})
    return cl


def _stat(html, label):
    """取某张数字卡上的数。数字紧挨在标签前，一起匹配才不会取到别的卡。"""
    m = re.search(r'>(\d+)</div>\s*<small class="text-muted">' + re.escape(label), html)
    assert m, f"首页上找不到「{label}」这张卡"
    return int(m.group(1))


# ---------------------------------------------------------------------------
# 四档按「本」算
# ---------------------------------------------------------------------------
def test_four_buckets_count_documents_not_applications(c):
    """按本算，不按出国申请条数算。

    甲持 3 本借走 1 本：在库该有他的港澳 + 台湾两本，不是「整个人算领用中」。
    乙没提过申请，他那本证按申请算根本进不了统计，按本算就在库里。
    """
    html = c.get("/").get_data(as_text=True)
    assert _stat(html, "在库（本）") == 3, "在库应为 甲的港澳 + 甲的台湾 + 乙的护照"
    assert _stat(html, "借出未还（本）") == 1, "借出的只有甲那本护照"
    assert _stat(html, "新办未入库（本）") == 2, "丁与戊都走路径B，新证都还没进台账"
    assert _stat(html, "逾期未交回（本）") == 2, "甲（借出超期）+ 丁（做证未交回超期），戊还没到期"


def test_stock_plus_lent_equals_ledger(c):
    """恒等式一：在库 + 借出未还 = 在控人员台账登记的总本数。

    这条恒等式是「在库」这个数能拿去和柜子对账的基础——两边对不上，
    说明有证既不在库也没借出，那是账本身错了。
    """
    html = c.get("/").get_data(as_text=True)
    total = _stat(html, "在库（本）") + _stat(html, "借出未还（本）")
    db = sqlite3.connect(Config.DATABASE)
    ledger = sum(r[0] for r in db.execute(
        "SELECT (CASE WHEN COALESCE(c.passport_no,'')<>'' THEN 1 ELSE 0 END)"
        "     + (CASE WHEN COALESCE(c.hm_pass_no ,'')<>'' THEN 1 ELSE 0 END)"
        "     + (CASE WHEN COALESCE(c.tw_pass_no ,'')<>'' THEN 1 ELSE 0 END) "
        "FROM certificates c JOIN personnel_filing pf ON pf.id=c.personnel_filing_id "
        "WHERE pf.status='active'"))
    db.close()
    assert total == ledger == 4, f"在库+借出={total}，台账={ledger}，对不上"


def test_overdue_is_a_subset_not_a_bigger_number(c):
    """恒等式二：逾期 ⊆ 借出未还 + 新办未入库。

    原来「逾期」会大于「领用中」——路径B 那条卡在「在库」桶里，进得了逾期
    进不了领用中。首页上摆着一个逻辑上不可能的数字组合，最需要被信任的告警
    反而最先让人怀疑系统算得对不对。
    """
    html = c.get("/").get_data(as_text=True)
    out = _stat(html, "借出未还（本）") + _stat(html, "新办未入库（本）")
    assert _stat(html, "逾期未交回（本）") <= out, "逾期比「在外」的总数还大"


def test_new_making_is_not_automatically_overdue(c):
    """「新办未入库」是一批证的去向，不是告警——里面只有超期的那部分才算逾期。

    戊下个月才出发，证在他手上是正常的，不该催；丁回国 120 天了还没交回，那才该催。
    两者都在「新办未入库」里，只有丁进「逾期未交回」。

    这条也钉住了恒等式里的 ⊆ 是**真子集**而不是恒等：少了戊这个样本，
    逾期恰好等于「借出 + 新办」，代码把 ⊆ 写成 = 也测不出来。
    """
    html = c.get("/").get_data(as_text=True)
    out = _stat(html, "借出未还（本）") + _stat(html, "新办未入库（本）")
    assert out == 3 and _stat(html, "逾期未交回（本）") == 2, \
        "戊（还没出发）被算进逾期了，或者丁（早该交回）没被算进去"


def test_decontrolled_certificate_is_not_in_stock(c):
    """已撤控人员的证不算在库：撤控以证件收缴移交为前提，那本证不在柜子里。

    台账行还留着是为了留痕（第 5 批 B1 给它标了「已撤控 · 移交」），
    不是因为证还在。
    """
    assert _cell("E3") not in c.get("/certificate/stock").get_data(as_text=True)


def test_monthly_issuance_card_is_gone(c):
    """「本月证件领用」删掉：那是工作量指标，不是待办，看了不产生任何动作。"""
    assert "本月证件领用" not in c.get("/").get_data(as_text=True)


def test_dashboard_computes_nothing_it_does_not_render(c):
    """算了不渲染的数，等于每进一次首页白跑一次查询。

    断言模板上下文而不是页面文字：这些变量本来就不渲染，只看页面永远是绿的。
    """
    from flask import template_rendered
    from app import create_app

    app = create_app()
    seen = {}
    def record(sender, template, context, **extra):
        if template.name == "dashboard.html":
            seen.update(context)
    template_rendered.connect(record, app)
    try:
        cl = app.test_client()
        tok = _CSRF.search(cl.get("/login").get_data(as_text=True)).group(1)
        cl.post("/login", data={"username": "admin", "password": "admin123", "csrf_token": tok})
        cl.get("/")
    finally:
        template_rendered.disconnect(record, app)

    assert seen, "没抓到首页的模板上下文，这条用例什么也没验证"
    for dead in ("expiring", "warn_days", "overdue", "iss_this_month", "iss_pending",
                 "cert_in_storage", "cert_in_use", "by_unit", "by_political", "by_rank"):
        assert dead not in seen, f"{dead} 仍在算并传给模板，而模板不用它"


# ---------------------------------------------------------------------------
# 盘库清单
# ---------------------------------------------------------------------------
def test_stock_list_names_every_document(c):
    """首页那个数只能核对总数；盘库清单要能逐本对，所以必须一本一行。"""
    html = c.get("/certificate/stock").get_data(as_text=True)
    for no in ("E2", "C1", "T1"):
        assert _cell(no) in html, f"应在库的 {no} 没列出来"
    assert _cell("E1") in html, "借出未还的那本也要列，否则对不上时不知道去哪儿了"


def _cell(cert_no):
    """证件号码在盘库清单里的渲染形态：<code>E1</code>。

    不能拿裸的 "E1" 去 html.find()——页面上还有一个**随机 CSRF 令牌**
    （token_urlsafe(32)，43 个字符），它有约 1% 的概率恰好包含 "E1" 这样的
    两字符片段。撞上了，find() 命中的就是令牌里那两个字符，与表格毫无关系，
    断言随机翻车。这条不是假想：把 bcrypt 成本调低重跑整套时真的撞到过一次。
    """
    return f"<code>{cert_no}</code>"


def _stock_row(html, cert_no):
    """截出盘库清单里含该证件号码的那一行 <tr>。

    整页断言分不清「在库」出现在哪一行——页面上到处都是这两个词（筛选下拉、
    说明、统计徽章）。要看的是**这本证那一行的去向列**。
    """
    i = html.find(_cell(cert_no))
    assert i != -1, f"清单里找不到 {cert_no}"
    start = html.rfind("<tr", 0, i)
    end = html.find("</tr>", i)
    assert start != -1 and end != -1, f"{cert_no} 不在任何表格行里"
    return html[start:end]


def test_stock_list_marks_where_each_document_is(c):
    """一张表 + 「去向」列：柜子里对不上的，答案在同一张纸的同一行上。

    原来分成「应在库」「借出未还」两张表，于是通用的那套列表行为（勾选、排序、
    窗口化分页、批量打印）每样都得写两遍，而且没有哪一份是「整份清单」。
    """
    html = c.get("/certificate/stock").get_data(as_text=True)
    assert "借出未还" in _stock_row(html, "E1"), "借出的那本没标成「借出未还」"
    for no in ("C1", "T1", "E2"):
        assert "在库" in _stock_row(html, no), f"{no} 没标成「在库」"
        assert "借出未还" not in _stock_row(html, no), f"{no} 被标成了「借出未还」"


def test_stock_list_can_be_filtered(c):
    """500 人的柜子，按种类、姓名或去向分批盘是常态。"""
    only_hm = c.get("/certificate/stock?cert_type=往来港澳通行证").get_data(as_text=True)
    assert _cell("C1") in only_hm and _cell("E2") not in only_hm

    only_yi = c.get("/certificate/stock?search=乙").get_data(as_text=True)
    assert _cell("E2") in only_yi and _cell("C1") not in only_yi

    only_out = c.get("/certificate/stock?status=借出未还").get_data(as_text=True)
    assert _cell("E1") in only_out, "「借出未还」筛不出借出的那本"
    assert _cell("E2") not in only_out and _cell("C1") not in only_out, \
        "「借出未还」把在库的也筛出来了"


def test_stock_page_and_dashboard_agree(c):
    """两处必须同口径——它们本来就该是同一个函数算出来的。"""
    dash = c.get("/").get_data(as_text=True)
    stock = c.get("/certificate/stock").get_data(as_text=True)
    m = re.search(r'badge bg-success">在库 (\d+) 本', stock)
    assert m, "盘库页上没有在库的计数"
    assert int(m.group(1)) == _stat(dash, "在库（本）")


def test_stock_export_lists_the_same_documents(c):
    """导出的就是屏幕上那份，包括筛选条件。"""
    r = c.get("/export/cert-stock")
    assert r.status_code == 200
    assert r.data[:2] == b"PK", "不是 xlsx（zip 容器）"

    import io
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(r.data)).active
    vals = {str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None}
    for no in ("E1", "E2", "C1", "T1"):
        assert no in vals, f"导出里缺 {no}"
    assert "E3" not in vals, "已撤控人员的证不该出现在盘库清单里"


def test_stock_export_respects_the_filter(c):
    import io
    from openpyxl import load_workbook
    r = c.get("/export/cert-stock?cert_type=往来港澳通行证")
    ws = load_workbook(io.BytesIO(r.data)).active
    vals = {str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None}
    assert "C1" in vals and "E2" not in vals


def test_orphan_issuance_number_is_surfaced(c):
    """借出记录上的号码在台账里找不到——不影响在库数，但说明账对不上，要报出来。"""
    db = sqlite3.connect(Config.DATABASE)
    db.execute("UPDATE cert_issuance SET cert_nos='E-NOT-IN-LEDGER' WHERE id=1")
    db.commit(); db.close()

    for page in ("/", "/certificate/stock"):
        html = c.get(page).get_data(as_text=True)
        assert "E-NOT-IN-LEDGER" in html, f"{page} 没有报出对不上的借出号码"
    # 台账上甲那本护照此时没人借，应当回到「应在库」
    assert c.get("/").get_data(as_text=True).count("E-NOT-IN-LEDGER") >= 1
    assert _stat(c.get("/").get_data(as_text=True), "在库（本）") == 4


# ---------------------------------------------------------------------------
# 盘库清单：与全站列表同形（勾选 / 排序 / 窗口化分页 / 批量打印）
# ---------------------------------------------------------------------------
def test_stock_list_uses_the_standard_list_markup(c):
    """盘库清单必须长成全站列表的样子，否则通用行为一样也用不上。

    此前它是两张自造的 table-sm 表：没有 id="mainTable"，于是 main.js 里那套
    窗口化分页与表头排序整个不生效（那段代码只认这一个 id）；没有 .row-check，
    于是勾选、导出选中行、批量打印也都没有。57 本证一次全倒在一页上。
    """
    html = c.get("/certificate/stock").get_data(as_text=True)
    assert html.count('id="mainTable"') == 1, \
        "盘库清单没有唯一的 #mainTable —— 窗口化分页与表头排序只认这个 id"
    assert 'id="selectAll"' in html, "缺少全选框"
    assert 'class="row-check"' in html, "缺少行勾选框（导出选中行 / 打印选中行都要它）"
    assert "table-sm" not in html, \
        "盘库清单用了 table-sm，行高与其余列表不一致"


def test_stock_rows_carry_a_stable_key_not_the_cert_number(c):
    """勾选框的值是「台账行 id + 号码槽」，不是证件号码。

    号码本该唯一，但数据出错时会重复；用号码当 key，勾一行会连带勾中另一个人的证。
    """
    html = c.get("/certificate/stock").get_data(as_text=True)
    assert 'value="1:passport_no"' in html
    assert 'value="1:hm_pass_no"' in html


def test_stock_print_page_is_standalone(c):
    """打印走独立排版页，不是把整张网页打出来。

    整页打印会把侧边栏、筛选表单、分页条一并印上纸；更糟的是窗口化分页只显示
    当前页，打出来的清单是残的。
    """
    html = c.get("/certificate/stock/print").get_data(as_text=True)
    assert "因私出国（境）证件盘库清单" in html
    assert "sidebar" not in html and 'id="mainTable"' not in html, "打印页混进了主界面结构"
    assert "display: table-header-group" in html, "表头没有设置跨页重复"
    assert "盘点人（签字）" in html, "盘点表没有签字栏"
    for no in ("E1", "E2", "C1", "T1"):
        assert no in html, f"打印页缺 {no}"


def test_stock_print_selected_only(c):
    """打印选中行：勾了哪几本就只印哪几本。"""
    html = c.get("/certificate/stock/print?ids=2:passport_no").get_data(as_text=True)
    assert "E2" in html
    for no in ("E1", "C1", "T1"):
        assert no not in html, f"只勾了 E2，{no} 也被印出来了"


def test_stock_selected_ids_win_over_other_filters(c):
    """勾选行优先：勾了就按勾的来，其余筛选不再叠加。

    否则「勾了 3 行却导出 2 行」说不清是谁的问题。
    """
    html = c.get("/certificate/stock/print?ids=1:passport_no&cert_type=往来港澳通行证"
                 ).get_data(as_text=True)
    assert "E1" in html, "勾中的 E1 被 cert_type 筛掉了"
    assert "C1" not in html


def test_stock_export_notes_are_one_per_line(c):
    """填表说明一条一行。

    NOTES_STOCK 曾经是一对括号包着的隐式拼接**字符串**（少了逗号），
    而 _save_and_return 里 `for note in notes` 逐**字符**迭代 ——
    打出来的填表说明一个字一行，几十行才拼出一句话。
    """
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(c.get("/export/cert-stock").data))
    ws = wb["填表说明"]
    lines = [r[0] for r in ws.iter_rows(values_only=True) if r[0]]
    assert 2 <= len(lines) <= 20, f"填表说明有 {len(lines)} 行，不像是一条一行"
    assert all(len(str(l)) > 1 for l in lines), \
        f"填表说明被逐字符拆开了：{lines[:5]}"
    assert any("在库" in str(l) for l in lines), "填表说明里没讲清口径"
