"""P1+P2 安全与健壮性加固测试：登录防爆破 / PDF 魔数 / 索引 / 错误页。"""
import io
import re
import sqlite3

import pytest

from config import Config

CSRF = re.compile(r'name="csrf-token" content="([^"]+)"')


@pytest.fixture()
def app_client(tmp_path, monkeypatch):
    monkeypatch.setattr(Config, "DATABASE", str(tmp_path / "t.db"))
    up = tmp_path / "up"
    up.mkdir()
    monkeypatch.setattr(Config, "UPLOAD_FOLDER", str(up))
    monkeypatch.setattr(Config, "EXPORT_FOLDER", str(tmp_path / "exp"))
    import database
    database.init_db()
    database.run_migrations()
    database.seed_data()
    import auth
    auth._login_fails.clear()  # 防爆破计数为进程级，测试间必须清零
    from app import create_app
    app = create_app()
    return app.test_client()


def _tok(c, path="/login"):
    return CSRF.search(c.get(path).get_data(as_text=True)).group(1)


def _login(c, pw, username="admin"):
    return c.post("/login", data={"username": username, "password": pw,
                                  "csrf_token": _tok(c)})


# ------------------------- S1 登录防爆破 -------------------------
def test_lockout_after_5_failures(app_client):
    c = app_client
    for _ in range(5):
        r = _login(c, "wrong-password")
        assert r.status_code == 200
    # 已锁定：即使密码正确也拒绝
    r = _login(c, "admin123")
    html = r.get_data(as_text=True)
    assert "锁定" in html
    assert r.status_code == 200  # 未跳转仪表盘
    # 确认未登录
    assert c.get("/").status_code == 302


def test_success_resets_counter(app_client):
    c = app_client
    for _ in range(3):
        _login(c, "wrong-password")
    r = _login(c, "admin123")
    assert r.status_code == 302  # 成功跳转
    # 登出后失败计数应已清零：再错 4 次仍未锁
    c.get("/logout")
    for _ in range(4):
        _login(c, "wrong-password")
    r = _login(c, "admin123")
    # 第 5 次失败才锁；上面只错了 4 次，本次正确应成功
    assert r.status_code == 302


def test_lock_event_logged(app_client):
    c = app_client
    for _ in range(5):
        _login(c, "wrong-password")
    db = sqlite3.connect(Config.DATABASE)
    row = db.execute("SELECT action, detail FROM operation_logs WHERE action='lock'").fetchone()
    assert row is not None and "锁定" in row[1]


# ------------------------- S2 PDF 魔数校验 -------------------------
def _travel_form(csrf, fake=False):
    pdf = b"NOT A PDF!" if fake else b"%PDF-1.4 fake body"
    return dict(
        csrf_token=csrf, personnel_filing_id="1", unit="局", department="科",
        name="张三", position="科员", title="工程师", id_number="110101199001012133",
        destination_passport="美国-护照", category="出国",
        travel_dates="2026/08/01-2026/08/11", need_new_passport="否",
        approval_date="20260701",
        passport_collect_date="20260725",
        att_application=(io.BytesIO(pdf), "a.pdf"),
        att_approval=(io.BytesIO(b"%PDF-1.4"), "b.pdf"),
    )


@pytest.fixture()
def logged_in(app_client):
    c = app_client
    db = sqlite3.connect(Config.DATABASE)
    db.execute("INSERT INTO personnel_filing (id,surname,given_name,gender,birth_date,id_number,"
               "residence,political_status,work_unit,position_or_title,supervisor_unit,tag,informed,"
               "status,operator) VALUES (1,'张','三','男','19900101','110101199001012133','X','群众',"
               "'局','科员','主管','新增','否','active','admin')")
    # 名下有一本在有效期内的护照：出国明细里「是否做证＝否」要求确有可用证件，
    # 否则那条记录本身就是错的（人没有证却说不用做证）。
    db.execute("INSERT INTO certificates (personnel_filing_id,unit,department,name,"
               "passport_no,passport_expiry,passport_submit_date,operator) "
               "VALUES (1,'局','科','张三','E12345678','20351231','20250101','admin')")
    db.commit()
    db.close()
    r = _login(c, "admin123")
    assert r.status_code == 302
    return c

def test_fake_pdf_rejected(logged_in):
    c = logged_in
    r = c.post("/travel/new", data=_travel_form(_tok(c, "/"), fake=True),
               content_type="multipart/form-data", follow_redirects=True)
    assert "不是有效的 PDF" in r.get_data(as_text=True)
    db = sqlite3.connect(Config.DATABASE)
    assert db.execute("SELECT COUNT(*) FROM travel_details").fetchone()[0] == 0  # 记录未入库


def test_real_pdf_accepted(logged_in):
    c = logged_in
    r = c.post("/travel/new", data=_travel_form(_tok(c, "/")),
               content_type="multipart/form-data", follow_redirects=True)
    assert "已保存" in r.get_data(as_text=True)
    db = sqlite3.connect(Config.DATABASE)
    assert db.execute("SELECT COUNT(*) FROM attachments").fetchone()[0] == 2


# ------------------------- F1 数据库索引 -------------------------
def test_indexes_created(app_client):
    db = sqlite3.connect(Config.DATABASE)
    names = {r[0] for r in db.execute(
        "SELECT name FROM sqlite_master WHERE type='index' AND name LIKE 'idx_%'").fetchall()}
    assert {"idx_pf_id_number", "idx_pf_status", "idx_td_pf_id",
            "idx_cert_pf_id", "idx_dec_pf_id", "idx_att_travel_id",
            "idx_logs_created_at"} <= names


# ------------------------- R1 中文错误页 -------------------------
def test_404_chinese_page(app_client):
    r = app_client.get("/no-such-page-xyz")
    assert r.status_code == 404
    assert "页面不存在" in r.get_data(as_text=True)


# ------------------------- P3: 配置 / 备份标记 / 日志归档 / 全局搜索 -------------------------
def test_session_cookie_flags():
    assert Config.SESSION_COOKIE_HTTPONLY is True
    assert Config.SESSION_COOKIE_SAMESITE == "Lax"


def test_backup_daily_marker(tmp_path, monkeypatch):
    monkeypatch.setattr(Config, "DATABASE", str(tmp_path / "d.db"))
    monkeypatch.setattr(Config, "BACKUP_FOLDER", str(tmp_path / "bk"))
    (tmp_path / "d.db").write_bytes(b"x")
    import utils.backup as bk
    bk._checked_date = None
    r1 = bk.run_daily_backup()
    assert r1["created"] is True
    assert bk._checked_date == r1["date"]          # 当日标记已置
    r2 = bk.run_daily_backup()                     # 同日第二次：直接跳过
    assert r2["created"] is False and r2["path"] is None
    r3 = bk.run_daily_backup(force=True)           # force 不受标记影响
    assert r3["created"] is True


def test_change_snapshots_never_overwrite(tmp_path, monkeypatch):
    """改前快照从不覆盖——同一秒内连着改两次，两份都要留住。

    这份文件的全部价值就在于它是「那一次改动之前」的样子。每日备份一天一个文件名，
    第二次批量改动会把它盖掉，第一次改错了就再也退不回去——快照不能重蹈覆辙。
    """
    monkeypatch.setattr(Config, "DATABASE", str(tmp_path / "d.db"))
    monkeypatch.setattr(Config, "BACKUP_FOLDER", str(tmp_path / "bk"))
    (tmp_path / "d.db").write_bytes(b"first")
    import utils.backup as bk

    a = bk.snapshot_before_change("org_rename")
    (tmp_path / "d.db").write_bytes(b"second")
    b = bk.snapshot_before_change("org_rename")     # 同一秒、同一 tag

    assert a != b, f"同 tag 连着两次留下了同名文件，后一份盖掉了前一份：{a}"
    bak = tmp_path / "bk"
    assert (bak / a).read_bytes() == b"first", "第一份快照的内容被后来的改动覆盖了"
    assert (bak / b).read_bytes() == b"second"
    assert a.startswith("before_org_rename_") and a.endswith(".db"), \
        f"快照名看不出是哪次改动：{a}"


def test_change_snapshots_are_pruned_with_the_dailies(tmp_path, monkeypatch):
    """快照同样受 30 天保留期约束，否则备份目录只进不出。"""
    monkeypatch.setattr(Config, "DATABASE", str(tmp_path / "d.db"))
    monkeypatch.setattr(Config, "BACKUP_FOLDER", str(tmp_path / "bk"))
    bak = tmp_path / "bk"; bak.mkdir()
    (bak / "before_org_rename_20200101_101010.db").write_bytes(b"old")
    (bak / "before_org_rename_20991231_101010.db").write_bytes(b"new")
    # encoding 必须显式给：不给就用平台默认编码，Linux 上是 UTF-8（能写），
    # Windows 运行器上是 cp1252（编不了中文，直接 UnicodeEncodeError）。
    # CI 打的是 Windows 包，本地是 Linux，这种差异只会在 CI 上炸。
    (bak / "notes.txt").write_text("不是备份文件，别碰", encoding="utf-8")

    import utils.backup as bk
    assert bk.prune_old_backups() == 1
    assert not (bak / "before_org_rename_20200101_101010.db").exists()
    assert (bak / "before_org_rename_20991231_101010.db").exists()
    assert (bak / "notes.txt").exists(), "清理动了不属于备份的文件"


def test_logs_export_by_year(logged_in):
    c = logged_in
    # 上面 fixture 的登录/建档已产生日志；取当前本地年份导出
    from datetime import datetime
    year = datetime.now().strftime("%Y")
    r = c.get(f"/logs/export?year={year}")
    assert r.status_code == 200
    assert r.data[:2] == b"PK"                     # xlsx 是 zip 容器
    # 无效年份回列表页
    assert c.get("/logs/export?year=abc").status_code == 302


def test_global_search(logged_in):
    c = logged_in
    r = c.get("/search?q=张三")
    html = r.get_data(as_text=True)
    assert r.status_code == 200
    assert "人员备案" in html and "张三" in html
    # 空关键词提示页
    assert "一次搜遍" in c.get("/search").get_data(as_text=True)
    # 无结果
    assert "未找到" in c.get("/search?q=不存在的名字XYZ").get_data(as_text=True)


# ------------------------- 导入模板去除“操作人”列 + 自动写入 -------------------------
def test_import_template_no_operator_column(app_client):
    from app import create_app  # 确保应用上下文可用
    import utils.excel_import as ei
    from openpyxl import load_workbook
    buf = ei.generate_import_template()
    hdr = [c.value for c in load_workbook(buf).active[1]]
    assert "操作人" not in hdr
    assert len(hdr) == 20 and hdr[-1] == "备注"


def test_import_operator_from_session(app_client):
    import utils.excel_import as ei
    from openpyxl import load_workbook
    from app import create_app
    buf = io.BytesIO()
    load_workbook(ei.generate_import_template()).save(buf)  # 含自带示例行
    buf.seek(0)
    with create_app().app_context():                        # 解析需应用上下文（get_db）
        res = ei.parse_import_file(buf, operator="wangwu")
    assert res["success"] == 1
    row = sqlite3.connect(Config.DATABASE).execute(
        "SELECT operator FROM personnel_filing").fetchone()
    assert row[0] == "wangwu"      # 操作人来自会话，而非表格


def test_banner_survives_non_utf8_stdout(tmp_path):
    """启动横幅里的中文，在非 UTF-8 的 stdout 上不能把进程打死。

    Windows 上 stdout 被重定向（写日志文件、注册成服务）时，Python 改用系统
    ANSI 代码页；英文 Windows 是 cp1252，编不出中文，print 抛 UnicodeEncodeError，
    程序启动即崩。CI 的 exe 冒烟就是这么红的一次。

    这里用 PYTHONIOENCODING=cp1252 在子进程里复现同一条路径——不是模拟，
    是让解释器真的换成那个编码器。去掉 _force_utf8_console() 这条会红。
    """
    import os
    import subprocess
    import sys as _sys

    repo = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    code = (
        "import sys; sys.path.insert(0, r'''%s''');"
        "import app; app._force_utf8_console();"
        "print('  因私出国（境）人员审批管理系统')" % repo
    )
    env = {**os.environ, "PYTHONIOENCODING": "cp1252",
           "POTMS_BASE": str(tmp_path / "base")}   # 别在仓库里落下 .secret_key
    r = subprocess.run([_sys.executable, "-c", code], env=env,
                       capture_output=True)
    assert r.returncode == 0, r.stderr.decode("utf-8", "replace")
    assert "因私出国" in r.stdout.decode("utf-8", "replace")
