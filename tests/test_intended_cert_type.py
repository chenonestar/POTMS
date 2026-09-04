"""第 7 批：出国申请上加「拟用证件种类」——把一个一直靠猜的答案变成录入项。

在此之前，「这趟要用哪种证」在系统里根本没有结构化的答案，只有「地点、证照」
那段自由文本。这一个缺口一路串出三个后果：

1. **判不了「够不够用」**。travel._validate_form 只能判「一本都没有」，
   它自己的注释就写着原因：「系统不知道这趟要用哪种证……有港澳通行证但要去美国
   这类情形只能靠经办人自己看」。
2. **领用时领哪本没人管**。领用表单上的证件种类是个三选一的自由单选，与这趟
   去哪儿毫无关联——去香港领出一本护照，系统一句话都不说（实测复现）。
3. **历史数据只能靠猜**。infer_cert_type 那套三级推断之所以存在，根子就在这里。

补上这一栏之后，判据有了源头：新增/编辑时必填，「说不做证」按**那一本**去查
有没有、在不在有效期内，领用登记时必须与它一致。

存量怎么办：迁移里直接复用 infer_cert_type 回填（号码匹配 → 「地点、证照」里的
证件名称 → 该人只登记了一种证），三条都不成立就留空标「待核实」——
**不替他猜一个**，给出错误答案比判不出更糟，这是那个函数当初的教训。
留空的那些不参与领用比对：不能拿一个系统自己都没判出来的答案去挡人。
"""
import io
import re
import sqlite3

import pytest

from config import Config
from conftest import seed_required_attachments, valid_id

_CSRF = re.compile(r'name="csrf-token" content="([^"]+)"')
_PNG = __import__("tests.test_issuance", fromlist=["_PNG_DATA_URL"])._PNG_DATA_URL
_PDF = b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\ntrailer\n<<>>\n%%EOF\n"


def _paths(tmp_path, monkeypatch):
    monkeypatch.setattr(Config, "DATABASE", str(tmp_path / "t.db"))
    up = tmp_path / "up"; up.mkdir()
    monkeypatch.setattr(Config, "UPLOAD_FOLDER", str(up))
    monkeypatch.setattr(Config, "EXPORT_FOLDER", str(tmp_path / "exp"))
    monkeypatch.setattr(Config, "BACKUP_FOLDER", str(tmp_path / "bak"))
    import database
    return database


def _client():
    from app import create_app
    cl = create_app().test_client()
    tok = _CSRF.search(cl.get("/login").get_data(as_text=True)).group(1)
    cl.post("/login", data={"username": "admin", "password": "admin123", "csrf_token": tok})
    return cl


def _tok(cl):
    return _CSRF.search(cl.get("/").get_data(as_text=True)).group(1)


def _one(sql, *params):
    db = sqlite3.connect(Config.DATABASE)
    row = db.execute(sql, params).fetchone()
    db.close()
    return row[0] if row else None


def _person(db, pid, nm, gn):
    db.execute("INSERT INTO personnel_filing (id,surname,given_name,gender,birth_date,id_number,"
               "residence,political_status,work_unit,position_or_title,supervisor_unit,operator) "
               "VALUES (?,?,?,'男','19900101',?,'浙江宁波市鄞州区','群众','总部','科长','人事处','admin')",
               (pid, nm, gn, valid_id(pid)))


# ===========================================================================
# 一、录入与校验
# ===========================================================================
@pytest.fixture()
def c(tmp_path, monkeypatch):
    """甲只有港澳通行证，乙名下什么证都没有。"""
    database = _paths(tmp_path, monkeypatch)
    database.init_db(); database.run_migrations(); database.seed_data()
    db = sqlite3.connect(Config.DATABASE)
    _person(db, 1, "甲", "一")
    db.execute("INSERT INTO certificates (personnel_filing_id,unit,department,name,"
               "hm_pass_no,hm_pass_expiry,hm_pass_submit_date,operator) "
               "VALUES (1,'总部','技术部','甲一','C1','20351231','20250101','admin')")
    _person(db, 2, "乙", "二")
    db.commit(); db.close()
    return _client()


def _new_travel(cl, pfid=1, nm="甲一", need="否", cert_type="02", **over):
    d = {"csrf_token": _tok(cl), "personnel_filing_id": str(pfid), "unit": "总部",
         "department": "技术部", "name": nm, "position": "科长",
         "id_number": valid_id(pfid), "destination_passport": "中国香港/往来港澳通行证",
         "category": "01", "travel_dates": "2026/11/01-2026/11/11",
         "need_new_passport": need, "approval_date": "20261001",
         "intended_cert_type": cert_type,
         "att_application": (io.BytesIO(_PDF), "a.pdf"),
         "att_approval": (io.BytesIO(_PDF), "b.pdf")}
    if need == "是":
        d["att_consent"] = (io.BytesIO(_PDF), "c.pdf")
    d.update(over)
    return cl.post("/travel/new", data=d, content_type="multipart/form-data",
                   follow_redirects=True)


def test_it_is_required_and_starred(c):
    """必填，且表单上标 *。"""
    r = _new_travel(c, cert_type="")
    assert "拟用证件种类" in r.get_data(as_text=True) and "必填" in r.get_data(as_text=True)
    assert _one("SELECT COUNT(*) FROM travel_details") == 0

    html = c.get("/travel/new").get_data(as_text=True)
    label = re.search(r'<label[^>]*>拟用证件种类</label>', html)
    assert label and "required" in label.group(0), "没标 *"


def test_an_unknown_code_is_refused(c):
    """只认字典里的代码——下拉框绕得过，POST 绕不过。"""
    r = _new_travel(c, cert_type="99")
    assert "无效的证件种类代码" in r.get_data(as_text=True)
    assert _one("SELECT COUNT(*) FROM travel_details") == 0


def test_it_is_saved_and_shown(c):
    """存下来，并在列表、详情、打印件上都显示成中文。"""
    _new_travel(c)
    assert _one("SELECT intended_cert_type FROM travel_details") == "02"
    for url in ("/travel/", "/travel/1", "/print/travel/1"):
        assert "往来港澳通行证" in c.get(url).get_data(as_text=True), f"{url} 上没显示"


# ---------------------------------------------------------------------------
# 「说不做证」现在精确到那一本
# ---------------------------------------------------------------------------
def test_holding_the_wrong_kind_is_caught(c):
    """甲只有港澳通行证，却说这趟用护照且不做证——挡下。

    这正是补这一栏的头号理由。原来的判据只看「一本都没有」，甲有港澳通行证，
    于是一路放行；现在判据精确到那一本。
    """
    r = _new_travel(c, cert_type="01")
    body = r.get_data(as_text=True)
    assert "没有在有效期内的普通护照" in body
    assert _one("SELECT COUNT(*) FROM travel_details") == 0


def test_holding_the_right_kind_passes(c):
    """同一个人改说用港澳通行证——他确实有，放行。

    没有这条对照，上面那条用「一律拦死」也能变绿。
    """
    _new_travel(c, cert_type="02")
    assert _one("SELECT COUNT(*) FROM travel_details") == 1


def test_an_expired_one_still_counts_as_none(c):
    """过期的那本等于没有——「有证」照旧要算有效期。"""
    db = sqlite3.connect(Config.DATABASE)
    db.execute("UPDATE certificates SET hm_pass_expiry='20200101'")
    db.commit(); db.close()
    assert "没有在有效期内的往来港澳通行证" in _new_travel(c, cert_type="02").get_data(as_text=True)


def test_path_b_is_not_asked_to_already_hold_it(c):
    """路径B 是去办新证的，当然还没有——不能拿「你没有这本证」去拦他。"""
    _new_travel(c, pfid=2, nm="乙二", need="是", cert_type="01")
    assert _one("SELECT COUNT(*) FROM travel_details") == 1


# ===========================================================================
# 二、领用时必须与申请一致
# ===========================================================================
@pytest.fixture()
def d(tmp_path, monkeypatch):
    """甲持护照与港澳通行证各一本；一条申请写明这趟用港澳通行证。"""
    database = _paths(tmp_path, monkeypatch)
    database.init_db(); database.run_migrations(); database.seed_data()
    db = sqlite3.connect(Config.DATABASE)
    _person(db, 1, "甲", "一")
    db.execute("INSERT INTO certificates (personnel_filing_id,unit,department,name,"
               "passport_no,passport_expiry,passport_submit_date,"
               "hm_pass_no,hm_pass_expiry,hm_pass_submit_date,operator) "
               "VALUES (1,'总部','技术部','甲一','E1','20351231','20250101',"
               "'C1','20351231','20250101','admin')")
    db.execute("INSERT INTO travel_details (id,personnel_filing_id,unit,department,name,position,"
               "id_number,destination_passport,intended_cert_type,category,travel_dates,"
               "travel_start,travel_end,need_new_passport,approval_date,operator) VALUES "
               "(1,1,'总部','技术部','甲一','科长',?,'中国香港/往来港澳通行证','02','01',"
               "'2026/11/01-2026/11/11','20261101','20261111','否','20261001','admin')",
               (valid_id(1),))
    seed_required_attachments(db, 1, "否")
    db.commit(); db.close()
    return _client()


def _issue(cl, cert_type, cert_no):
    return cl.post("/issuance/new", data={
        "csrf_token": _tok(cl), "travel_id": "1", "personnel_filing_id": "1",
        "holder_name": "甲一", "id_number": valid_id(1), "cert_types": cert_type,
        "cert_nos": cert_no, "issue_date": "20261025", "sign_png": _PNG,
    }, follow_redirects=True)


def test_issuing_the_wrong_kind_is_refused(d):
    """申请写的是港澳通行证，却来领护照——挡下，并把两边都说出来。

    这就是报出来的那个场景：不出国却领用护照。此前系统一句话都不说。
    """
    body = _issue(d, "01", "E1").get_data(as_text=True)
    assert "往来港澳通行证" in body and "普通护照" in body and "不一致" in body
    assert _one("SELECT COUNT(*) FROM cert_issuance") == 0


def test_issuing_the_matching_kind_goes_through(d):
    """领对了就放行——判据不能变成一道谁也过不去的墙。"""
    _issue(d, "02", "C1")
    assert _one("SELECT cert_types FROM cert_issuance") == "02"


def test_the_form_preselects_and_states_the_intended_kind(d):
    """领用表单按申请上写明的那一种预选，并把这条约束讲在明处。

    校验那一关反正要求一致，让人先选一遍再被打回来是纯粹的为难。
    """
    html = d.get("/issuance/new?travel_id=1").get_data(as_text=True)
    assert "拟用证件种类" in html and "往来港澳通行证" in html
    radio = re.search(r'<input[^>]*value="02"[^>]*>', html)
    assert radio and "checked" in radio.group(0), "没有按申请预选港澳通行证"


def test_a_legacy_application_without_it_is_not_blocked(d):
    """申请上这一栏是空的（历史数据回填判不出的那批）——不比对，照常放行。

    不能拿一个系统自己都没判出来的答案去挡人。表单上也要说清楚这一点，
    否则经办人看不出为什么这次不校验。
    """
    db = sqlite3.connect(Config.DATABASE)
    db.execute("UPDATE travel_details SET intended_cert_type=NULL WHERE id=1")
    db.commit(); db.close()

    html = d.get("/issuance/new?travel_id=1").get_data(as_text=True)
    assert "未填写拟用证件种类" in html, "没说明这次为什么不校验"

    _issue(d, "01", "E1")
    assert _one("SELECT COUNT(*) FROM cert_issuance") == 1


# ===========================================================================
# 三、存量回填
# ===========================================================================
@pytest.fixture()
def legacy(tmp_path, monkeypatch):
    """先建一个**没有这一列**的老库，塞进四条出行，再跑迁移。

    四条正好覆盖 infer_cert_type 的三级判据加一个判不出的：
      1 号码对上台账的港澳列        → 第①级
      2 「地点、证照」里写了「护照」  → 第②级
      3 该人只登记了一种证（台湾）    → 第③级
      4 三本证都有、没填号码、文字里也没证件名 → 判不出，留空
    """
    database = _paths(tmp_path, monkeypatch)
    database.init_db()
    db = sqlite3.connect(Config.DATABASE)
    # 把新列摘掉，模拟升级前的库
    db.execute("ALTER TABLE travel_details DROP COLUMN intended_cert_type")
    for pid, nm in ((1, "甲"), (2, "乙"), (3, "丙"), (4, "丁")):
        _person(db, pid, nm, "一")
    db.execute("INSERT INTO certificates (personnel_filing_id,unit,department,name,"
               "hm_pass_no,hm_pass_expiry,hm_pass_submit_date,operator) "
               "VALUES (1,'总部','技术部','甲一','C1','20351231','20250101','admin')")
    db.execute("INSERT INTO certificates (personnel_filing_id,unit,department,name,"
               "passport_no,passport_expiry,passport_submit_date,"
               "hm_pass_no,hm_pass_expiry,hm_pass_submit_date,operator) "
               "VALUES (2,'总部','技术部','乙一','E2','20351231','20250101',"
               "'C2','20351231','20250101','admin')")
    db.execute("INSERT INTO certificates (personnel_filing_id,unit,department,name,"
               "tw_pass_no,tw_pass_expiry,tw_pass_submit_date,operator) "
               "VALUES (3,'总部','技术部','丙一','T3','20351231','20250101','admin')")
    db.execute("INSERT INTO certificates (personnel_filing_id,unit,department,name,"
               "passport_no,passport_expiry,passport_submit_date,"
               "hm_pass_no,hm_pass_expiry,hm_pass_submit_date,"
               "tw_pass_no,tw_pass_expiry,tw_pass_submit_date,operator) "
               "VALUES (4,'总部','技术部','丁一','E4','20351231','20250101',"
               "'C4','20351231','20250101','T4','20351231','20250101','admin')")
    for tid, pfid, dest, no in ((1, 1, "中国香港", "C1"),
                                (2, 2, "美国/护照", ""),
                                (3, 3, "台北", ""),
                                (4, 4, "某地", "")):
        db.execute("INSERT INTO travel_details (id,personnel_filing_id,unit,department,name,"
                   "position,id_number,destination_passport,category,travel_dates,"
                   "need_new_passport,passport_no,operator) VALUES "
                   "(?,?,'总部','技术部','某人','科长',?,?,'01','历史','否',?,'admin')",
                   (tid, pfid, valid_id(pfid), dest, no))
    db.commit(); db.close()
    database.run_migrations(); database.seed_data()
    return _client()


def _intended(tid):
    return _one("SELECT intended_cert_type FROM travel_details WHERE id=?", tid)


def test_the_column_is_added_to_an_old_database(legacy):
    """老库升级后有这一列，系统起得来。"""
    db = sqlite3.connect(Config.DATABASE)
    cols = {r[1] for r in db.execute("PRAGMA table_info(travel_details)")}
    db.close()
    assert "intended_cert_type" in cols
    assert legacy.get("/travel/").status_code == 200


def test_backfill_uses_the_three_tier_inference(legacy):
    """三级判据逐级点名，判不出的留空。

    分四条断言而不是合成一条：哪一级踩空了要一眼看得见。
    """
    assert _intended(1) == "02", "①号码对上台账的港澳列"
    assert _intended(2) == "01", "②「地点、证照」里写了护照"
    assert _intended(3) == "03", "③该人只登记了台湾通行证一种"
    assert (_intended(4) or "") == "", "判不出就该留空，不能替他猜一个"


def test_the_unfilled_one_is_shown_as_pending_not_blank(legacy):
    """判不出的在界面上写「待核实」，不是一个空格子。

    空白会被当成漏渲染；写明待核实才是实情，也才收得了口。
    """
    body = legacy.get("/travel/").get_data(as_text=True)
    body = body[body.find("<tbody"):body.find("</tbody>")]
    assert "待核实" in body
    assert "待核实" in legacy.get("/travel/4").get_data(as_text=True)


def test_the_backfill_does_not_run_twice(legacy):
    """再跑一次迁移不会覆盖已有的值——幂等靠「列在不在」判断。"""
    db = sqlite3.connect(Config.DATABASE)
    db.execute("UPDATE travel_details SET intended_cert_type='03' WHERE id=1")
    db.commit(); db.close()

    import database
    database.run_migrations()
    assert _intended(1) == "03", "第二次迁移把人工订正过的值冲掉了"


# ===========================================================================
# 四、导出
# ===========================================================================
def test_the_export_carries_it(c):
    """导出件上要有这一列，判不出的写「待核实」。"""
    _new_travel(c)
    db = sqlite3.connect(Config.DATABASE)
    db.execute("INSERT INTO travel_details (id,personnel_filing_id,unit,department,name,position,"
               "id_number,destination_passport,category,travel_dates,need_new_passport,operator) "
               "VALUES (9,1,'总部','技术部','甲一','科长',?,'某地','01','历史','否','admin')",
               (valid_id(1),))
    db.commit(); db.close()

    from openpyxl import load_workbook
    from app import create_app
    with create_app().app_context():
        from utils.excel_export import export_travel_details
        path, _ = export_travel_details("admin", "", ())
    ws = load_workbook(path).active
    header = [c.value for c in ws[2]]
    assert "拟用证件种类" in header
    col = header.index("拟用证件种类") + 1
    got = {ws.cell(row=r, column=col).value for r in range(3, 3 + ws.max_row)}
    assert "往来港澳通行证" in got
    assert "待核实" in got, "空值在导出件上是个空格子，看的人分不清是没有还是漏填"
