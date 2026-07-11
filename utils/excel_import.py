"""Excel 批量导入 — 逐行校验 + 错误报告"""
import io
from datetime import datetime

from openpyxl import load_workbook

from database import get_db
from utils.validators import validate_id_number, validate_birth_date_match, validate_date_format, parse_date_input
from utils.helpers import normalize_residence, detect_surname_split


def parse_import_file(file_stream, operator: str = "admin") -> dict:
    """
    解析上传的 Excel 文件，逐行校验，返回导入结果。
    operator 由当前登录会话自动传入（与其它录入入口一致，不需在表格中填写）。
    返回: { total, success, errors: [{row, field, message}], imported_ids: [...] }
    """
    wb = load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头

    result = {"total": len(rows), "success": 0, "errors": [], "imported_ids": []}
    db = get_db()

    for i, row in enumerate(rows, start=2):  # 第2行开始
        # 跳过完全空行
        if not row or all(v is None or str(v).strip() == "" for v in row):
            result["total"] -= 1
            continue

        try:
            data = _parse_row(row)
        except Exception as e:
            result["errors"].append({"row": i, "field": "—", "message": f"数据解析失败: {e}"})
            continue

        row_errors = _validate_import_row(data, db)
        if row_errors:
            for err in row_errors:
                result["errors"].append({"row": i, "field": err[0], "message": err[1]})
            continue

        try:
            # 写入 personnel_info
            db.execute(
                "INSERT INTO personnel_info (unit, department, name, gender, birth_date, "
                "id_number, work_start_date, education, degree, title, rank, political_status, "
                "party_join_date, position, operator) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    data["unit"], data["department"], data["name"], data["gender"],
                    data["birth_date"], data["id_number"], data["work_start_date"], data["education_code"],
                    data["degree_code"], data["title_code"], data["rank_code"],
                    data["political_status"], data["party_join_date"],
                    data["position"], operator,
                ),
            )
            info_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

            # 写入 personnel_filing
            surname, given_name = detect_surname_split(data["name"])
            db.execute(
                "INSERT INTO personnel_filing (personnel_info_id, surname, given_name, gender, "
                "birth_date, id_number, residence, political_status, work_unit, "
                "position_or_title, supervisor_unit, tag, informed, remarks, operator) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    info_id, surname, given_name, data["gender"],
                    data["birth_date"], data["id_number"], normalize_residence(data.get("residence", "")),
                    data["political_status"], data["unit"], data.get("position_or_title", ""),
                    data.get("supervisor_unit", "人事处"), "新增", data.get("informed", "是"),
                    data.get("remarks", ""), operator,
                ),
            )

            result["success"] += 1
            result["imported_ids"].append(info_id)
        except Exception as e:
            result["errors"].append({"row": i, "field": "—", "message": f"数据库写入失败: {e}"})

    db.commit()
    return result


def generate_import_template() -> io.BytesIO:
    """生成导入模板 Excel 文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "备案人员导入模板"

    headers = [
        "单位", "部门", "姓名", "性别", "出生日期", "参加工作日期",
        "身份证号", "户口所在地", "政治面貌", "职务（级）或职称",
        "人事主管单位", "学历", "学位", "职称", "职级",
        "入党日期", "职务（岗位名称）", "标记", "已告知本人", "备注",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3A5A7C", end_color="3A5A7C", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # 示例行
    example = [
        "XX单位", "XX部门", "张三", "男", "19800103", "20000701",
        "330102198001031230", "浙江杭州市西湖区", "中共党员", "处级",
        "人事处", "大学本科", "学士", "副高", "处级",
        "20050701", "处长", "新增", "是", "",
    ]
    for col, val in enumerate(example, 1):
        ws.cell(row=2, column=col, value=val)

    # 列宽
    col_widths = [18, 14, 10, 6, 12, 12, 20, 22, 14, 18, 14, 12, 10, 10, 10, 12, 18, 8, 12, 20]
    for i, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
def _parse_row(row) -> dict:
    """将Excel行转为字典"""
    fields = [
        "unit", "department", "name", "gender", "birth_date", "work_start_date",
        "id_number", "residence", "political_status", "position_or_title",
        "supervisor_unit", "education_code", "degree_code", "title_code", "rank_code",
        "party_join_date", "position", "tag", "informed", "remarks",
    ]
    data = {}
    for i, field in enumerate(fields):
        val = row[i] if i < len(row) else None
        data[field] = str(val).strip() if val is not None else ""
    # 日期清洗
    data["birth_date"] = parse_date_input(data["birth_date"])
    data["work_start_date"] = parse_date_input(data["work_start_date"])
    data["party_join_date"] = parse_date_input(data["party_join_date"])
    return data


def _validate_import_row(data: dict, db) -> list[tuple[str, str]]:
    """返回 [(字段名, 错误信息), ...]"""
    errors = []
    required = [
        ("unit", "单位"), ("department", "部门"), ("name", "姓名"),
        ("gender", "性别"), ("birth_date", "出生日期"), ("id_number", "身份证号"),
        ("political_status", "政治面貌"), ("position", "职务（岗位名称）"),
    ]
    for field, label in required:
        if not data.get(field):
            errors.append((label, "必填项为空"))

    if data.get("birth_date"):
        ok, msg = validate_date_format(data["birth_date"])
        if not ok:
            errors.append(("出生日期", msg))

    if data.get("id_number"):
        ok, msg = validate_id_number(data["id_number"])
        if not ok:
            errors.append(("身份证号", msg))
        elif data.get("birth_date"):
            ok2, msg2 = validate_birth_date_match(data["id_number"], data["birth_date"])
            if not ok2:
                errors.append(("出生日期/身份证号", msg2))

    # 检查重复
    if data.get("id_number"):
        dup = db.execute(
            "SELECT id FROM personnel_filing WHERE id_number = ? AND status = 'active'",
            (data["id_number"],),
        ).fetchone()
        if dup:
            errors.append(("身份证号", "系统中已存在有效备案记录"))

    return errors
