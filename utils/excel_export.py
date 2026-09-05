"""Excel 导出 — 使用 openpyxl 生成 5 类表单"""
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db
from config import Config

# 通用样式
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=16)
TITLE_ALIGN = Alignment(horizontal="center", vertical="center")


def _style_header(ws, title: str, headers: list):
    """写入标题行（表名，合并居中）+ 列头行，并冻结。数据从第 3 行开始。"""
    col_count = len(headers)
    # 第1行：表名标题（跨列合并居中）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = TITLE_FONT
    tcell.alignment = TITLE_ALIGN
    ws.row_dimensions[1].height = 30
    # 第2行：列头
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    # 冻结标题 + 列头
    ws.freeze_panes = "A3"


def _style_data(ws, start_row: int, end_row: int, col_count: int):
    """给数据区加边框和对齐"""
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER


def _auto_width(ws, col_count: int, max_width: int = 40, min_row: int = 2):
    """自动列宽（默认从列头行起算，跳过合并的标题行）"""
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_row=min_row, min_col=col, max_col=col, values_only=True):
            for val in row:
                if val:
                    max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, max_width)


_EXPORT_RETENTION_DAYS = 7


def _prune_old_exports() -> None:
    """清理超过保留期的历史导出文件（导出目录只增不减会长期累积，且含敏感数据）。"""
    if not os.path.isdir(Config.EXPORT_FOLDER):
        return
    cutoff = datetime.now().timestamp() - _EXPORT_RETENTION_DAYS * 86400
    for name in os.listdir(Config.EXPORT_FOLDER):
        if not name.lower().endswith(".xlsx"):
            continue
        path = os.path.join(Config.EXPORT_FOLDER, name)
        try:
            if os.path.isfile(path) and os.path.getmtime(path) < cutoff:
                os.remove(path)
        except OSError:
            pass  # 单个文件清理失败不影响导出


def _save_and_return(ws, prefix: str, operator: str, notes: list = None):
    """保存到文件，返回路径"""
    # 添加填表说明 Sheet
    if notes:
        # 传成一整个字符串时按「一条说明」处理，而不是逐字符迭代。
        # 盘库清单的 NOTES 曾经就是这样一个字符串（隐式拼接少了逗号），
        # 打出来一个字一行——这行防御让同类笔误只影响排版，不再毁掉整张说明。
        if isinstance(notes, str):
            notes = [notes]
        ws2 = ws.parent.create_sheet("填表说明")
        # 说明是长句，默认列宽会把它挤成一列；给足宽度，靠左顶格。
        ws2.column_dimensions["A"].width = 96
        for i, note in enumerate(notes, 1):
            cell = ws2.cell(row=i, column=1, value=note)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_{ts}_{operator}.xlsx"
    filepath = os.path.join(Config.EXPORT_FOLDER, filename)
    os.makedirs(Config.EXPORT_FOLDER, exist_ok=True)
    _prune_old_exports()
    ws.parent.save(filepath)
    return filepath, filename


# =========================================================================
# 1. 备案人员信息登记表
# =========================================================================
HEADERS_INFO = [
    "单位", "部门", "姓名", "性别", "出生日期", "身份证号", "参加工作日期",
    "学历", "学位", "职称", "职级", "政治面貌", "入党日期", "职务（岗位名称）",
]

NOTES_INFO = [
    "填表说明：",
    "1. 出生日期格式为YYYYMMDD，需与身份证号对应。",
    "2. 学历、学位、职称、职级、政治面貌从系统数据字典中选择。",
    "3. 中共党员/预备党员须填写入党日期。",
]


def export_personnel_info(operator: str, where_sql: str = "", params: tuple = (), joined: bool = False) -> str:
    db = get_db()
    # #4 一律经 personnel_filing 关联导出：只导出有备案引用的信息登记表，
    # 无引用的孤儿行永不外泄（GROUP BY 去重，避免一人多条备案时重复）。
    # joined 参数保留以兼容旧调用，实际行为恒为关联导出。
    sql = ("SELECT pi.* FROM personnel_info pi "
           "JOIN personnel_filing pf ON pf.personnel_info_id = pi.id "
           "WHERE 1=1 " + where_sql + " GROUP BY pi.id ORDER BY pi.created_at DESC")
    rows = db.execute(sql, params).fetchall()

    # 学历/学位/职称/职级为字典编码，导出时映射为显示值（编码 → 中文）
    from utils.helpers import get_dict_options
    dict_maps = {
        cat: {o["code"]: o["value"] for o in get_dict_options(cat)}
        for cat in ("education", "degree", "title", "rank")
    }

    def _dv(cat, code):
        return dict_maps[cat].get(code, code) if code else ""

    wb = Workbook()
    ws = wb.active
    ws.title = "备案人员信息登记表"
    _style_header(ws, "备案人员信息登记表", HEADERS_INFO)

    for i, row in enumerate(rows, 3):
        values = [
            row["unit"], row["department"], row["name"], row["gender"],
            row["birth_date"], row["id_number"] or "", row["work_start_date"] or "",
            _dv("education", row["education"]), _dv("degree", row["degree"]),
            _dv("title", row["title"]), _dv("rank", row["rank"]),
            row["political_status"], row["party_join_date"] or "",
            row["position"],
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=i, column=col, value=val)

    _style_data(ws, 3, len(rows) + 2, len(HEADERS_INFO))
    _auto_width(ws, len(HEADERS_INFO))
    return _save_and_return(ws, "备案人员信息登记表", operator, NOTES_INFO)


# =========================================================================
# 2. 因私事出国（境）人员登记备案表
# =========================================================================
HEADERS_FILING = [
    "中文姓", "中文名", "性别", "出生日期", "身份证号", "户口所在地",
    "政治面貌", "工作单位", "职务（级）或职称", "人事主管单位",
    "标记", "已告知本人", "状态", "备注",
]

NOTES_FILING = [
    "填表说明：",
    "1. 姓与名分开填写，特别注意复姓人员。",
    "2. 出生日期格式为YYYYMMDD，生日需与身份证号对应。",
    "3. 工作单位请写全称。",
    "4. 职务/职称栏：处级领导填'处级'或'副处级'，副处级单位班子成员填'正科'，其他人员填'副高'或'正高'。",
    "5. 人事主管单位名称需与印章一致。",
    "6. 户口所在地填至区级，省份不加'省'字，江东区、鄞县统一为'鄞州区'。",
    "7. 标记：新增、更新。",
    "8. 已告知本人：是、否。",
]


def export_personnel_filing(operator: str, where_sql: str = "", params: tuple = ()) -> str:
    db = get_db()
    rows = db.execute(
        "SELECT pf.* FROM personnel_filing pf "
        "LEFT JOIN personnel_info pi ON pf.personnel_info_id = pi.id "
        "WHERE 1=1 " + where_sql + " ORDER BY pf.created_at DESC",
        params,
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "登记备案表"
    _style_header(ws, "因私事出国（境）人员登记备案表", HEADERS_FILING)

    for i, row in enumerate(rows, 3):
        values = [
            row["surname"], row["given_name"], row["gender"], row["birth_date"],
            row["id_number"], row["residence"], row["political_status"],
            row["work_unit"], row["position_or_title"], row["supervisor_unit"],
            row["tag"], row["informed"],
            "有效" if row["status"] == "active" else "已撤控",
            row["remarks"] or "",
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=i, column=col, value=val)

    _style_data(ws, 3, len(rows) + 2, len(HEADERS_FILING))
    _auto_width(ws, len(HEADERS_FILING))
    return _save_and_return(ws, "登记备案表", operator, NOTES_FILING)


# =========================================================================
# 3. 证照登记表
# =========================================================================
HEADERS_CERT = [
    "单位", "部门", "姓名",
    "普通护照", "护照证件号", "护照有效日期", "护照上交日期",
    "港澳通行证", "港澳通行证号", "港澳通有效日期", "港澳通上交日期",
    "台湾通行证", "台湾通行证号", "台湾通有效日期", "台湾通上交日期",
]

NOTES_CERT = [
    "填表说明：",
    "1. 一人可同时持有多类证件，无某类证件则留空。",
    "2. 填写某类证件号时，其有效日期与上交日期均为必填。",
    "3. 日期格式为 YYYYMMDD。",
    "4. 系统对有效期到期前 30 天进行预警提示。",
]


def export_certificates(operator: str, where_sql: str = "", params: tuple = ()) -> str:
    db = get_db()
    rows = db.execute(
        "SELECT * FROM certificates WHERE 1=1 " + where_sql + " ORDER BY updated_at DESC",
        params,
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "证照登记表"
    _style_header(ws, "因私出国（境）备案人员证照登记表", HEADERS_CERT)

    for i, row in enumerate(rows, 3):
        values = [
            row["unit"], row["department"], row["name"],
            "普通护照", row["passport_no"] or "", row["passport_expiry"] or "", row["passport_submit_date"] or "",
            "往来港澳通行证", row["hm_pass_no"] or "", row["hm_pass_expiry"] or "", row["hm_pass_submit_date"] or "",
            "大陆居民往来台湾通行证", row["tw_pass_no"] or "", row["tw_pass_expiry"] or "", row["tw_pass_submit_date"] or "",
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=i, column=col, value=val)

    _style_data(ws, 3, len(rows) + 2, len(HEADERS_CERT))
    _auto_width(ws, len(HEADERS_CERT))
    return _save_and_return(ws, "证照登记表", operator, NOTES_CERT)


# =========================================================================
# 3.5 证件盘库清单
# =========================================================================
HEADERS_STOCK = ["核对", "序号", "去向", "单位", "部门", "持证人",
                 "证件种类", "证件号码", "有效期至", "上交日期"]
# 必须是「多行」的序列，一行一条。
# 曾经写成一对括号包着的隐式拼接字符串（少了逗号），于是 _save_and_return 里
# for note in notes 逐**字符**迭代，填表说明打出来一个字一行。
NOTES_STOCK = [
    "填表说明：",
    "1. 本表列的是在控人员证照台账上登记的每一本证，一本一行，按「去向」分为在库与借出未还。",
    "2. 在库 = 此刻应当躺在保管处的证；借出未还 = 凭领用单借出、尚未归还的证。",
    "   两者相加等于在控人员台账登记的总本数。",
    "3. 已撤控人员的证不在此表（撤控以证件收缴移交为前提，证已随撤控移交出库）。",
    "4. 做证人员自办、尚未交回入库的新证也不在此表（它还没进过台账，也没进过柜子）。",
    "5. 「核对」列供盘库时逐本打勾。",
]


def export_cert_stock(operator: str, rows) -> str:
    """盘库清单导出。数据由调用方算好传入——口径只有 certificate.stock_rows 一处，
    这里再查一遍就等于开了第二套判据，页面上看到的和导出的迟早对不上。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "证件盘库清单"
    _style_header(ws, "因私出国（境）证件盘库清单", HEADERS_STOCK)

    for n, it in enumerate(rows, 1):
        for col, val in enumerate(
            ["", n, it["status"], it["unit"], it["department"], it["name"],
             it["cert_type"], it["cert_no"], it["expiry"], it["submit_date"]], 1):
            ws.cell(row=n + 2, column=col, value=val)

    _style_data(ws, 3, len(rows) + 2, len(HEADERS_STOCK))
    _auto_width(ws, len(HEADERS_STOCK))
    return _save_and_return(ws, "证件盘库清单", operator, NOTES_STOCK)


# =========================================================================
# 4. 因私出国（境）人员明细表
# =========================================================================
def _cert_type_label(code) -> str:
    """证件种类代码 → 中文；空值写「待核实」。

    导出件与打印件上不能是个空格子——看的人分不清是「没有」还是「漏填了」。
    与 issuance._types_label 同一条规约（那边的注释写明了理由）。
    """
    from utils.helpers import get_dict_value
    code = (code or "").strip()
    return (get_dict_value("cert_type", code) or code) if code else "待核实"


HEADERS_TRAVEL = [
    "单位", "部门", "姓名", "职务", "职称", "身份证号",
    "目的地", "拟用证件种类", "类别", "计划出行日期", "批准日期",
    "是否做证", "证件号码", "证件领用日期", "实际回国日期",
    "证件归还日期", "行程状态", "取消日期",
]


def export_travel_details(operator: str, where_sql: str = "", params: tuple = ()) -> str:
    db = get_db()
    rows = db.execute(
        "SELECT * FROM travel_details WHERE 1=1 " + where_sql + " ORDER BY created_at DESC",
        params,
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "出国明细表"
    _style_header(ws, "因私出国（境）人员明细表", HEADERS_TRAVEL)

    for i, row in enumerate(rows, 3):
        values = [
            row["unit"], row["department"], row["name"], row["position"],
            row["title"] or "", row["id_number"], row["destination_passport"],
            _cert_type_label(row["intended_cert_type"]),
            row["category"], row["travel_dates"], row["approval_date"] or "",
            row["need_new_passport"], row["passport_no"] or "",
            row["passport_collect_date"] or "", row["actual_return_date"] or "",
            row["passport_return_date"] or "",
            "取消行程" if row["trip_status"] == "cancelled" else "正常",
            row["cancel_date"] or "",
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=i, column=col, value=val)

    _style_data(ws, 3, len(rows) + 2, len(HEADERS_TRAVEL))
    _auto_width(ws, len(HEADERS_TRAVEL))
    return _save_and_return(ws, "出国明细表", operator, [
        "1. 计划出行日期格式：起始日期-结束日期，如 2023-6-20-2023-6-26。",
        "2. 附件需线下查看系统存储的PDF扫描件。",
    ])


# =========================================================================
# 5. 撤控备案表
# =========================================================================
HEADERS_DEC = [
    "中文姓", "中文名", "性别", "出生日期", "身份证号", "户口所在地",
    "政治面貌", "工作单位", "人事主管单位", "报送单位名称",
    "报送单位类别", "报送单位联系人", "报送单位联系电话",
    "入库批号", "撤控日期", "证件移交日期", "撤控原因",
]


def export_decontrol(operator: str, where_sql: str = "", params: tuple = ()) -> str:
    db = get_db()
    rows = db.execute(
        "SELECT * FROM decontrol_filing WHERE 1=1 " + where_sql + " ORDER BY created_at DESC",
        params,
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "撤控备案表"
    _style_header(ws, "因私事出国（境）人员撤控备案表", HEADERS_DEC)

    for i, row in enumerate(rows, 3):
        values = [
            row["surname"], row["given_name"], row["gender"], row["birth_date"],
            row["id_number"], row["residence"], row["political_status"],
            row["work_unit"], row["supervisor_unit"], row["submit_unit_name"],
            row["submit_unit_type"], row["submit_contact"], row["submit_phone"],
            row["batch_no"], row["decontrol_date"] or "", row["cert_handover_date"] or "", row["reason"],
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=i, column=col, value=val)

    _style_data(ws, 3, len(rows) + 2, len(HEADERS_DEC))
    _auto_width(ws, len(HEADERS_DEC))
    return _save_and_return(ws, "撤控备案表", operator, [
        "1. 出生日期格式为YYYYMMDD，生日需与身份证号对应。",
        "2. 户口所在地填至区级，省份不加'省'字。",
        "3. 报送单位类别：党政机关,金融系统,教科文卫系统,国有大中型企业单位,其他单位。",
    ])


# =========================================================================
# 6. 证件领用登记表（含手写签名图片）
# =========================================================================
HEADERS_ISS = ["单位", "领用人", "身份证号", "证件种类", "证件号码", "领用日期",
               "经办人(发放)", "领用签名", "归还日期", "经办人(接收)", "归还签名",
               "状态", "备注"]

_STATUS_LABEL = {"issued": "已领用", "returned": "已归还", "voided": "已作废"}

_PNG_MAGIC = b"\x89PNG\r\n\x1a\n"


def _png_size(data: bytes) -> tuple:
    """从 PNG 的 IHDR 块解析 (宽, 高)。

    PNG 结构：8 字节签名 + 4 字节块长度 + 'IHDR' + 宽(4) + 高(4)，
    故宽高固定位于第 16–24 字节，无需图像库即可读取。
    """
    if len(data) < 24 or data[:8] != _PNG_MAGIC or data[12:16] != b"IHDR":
        raise ValueError("不是有效的 PNG 数据")
    return (int.from_bytes(data[16:20], "big"), int.from_bytes(data[20:24], "big"))


def _make_png_image(data: bytes):
    """构造 openpyxl 图片对象（不依赖 Pillow）。

    openpyxl 的 Image 需要 Pillow 才能读出尺寸；但签名图固定为 PNG，
    尺寸可自行解析，故在此继承并跳过 Pillow 分支 —— 本项目依赖刻意保持
    精简（纯 Python，便于打包单文件 exe），不为读一个尺寸引入 C 扩展。
    继承是必需的：openpyxl 序列化时以 isinstance(obj, Image) 判定图片。
    """
    import io
    from openpyxl.drawing.image import Image as XLImage

    class _PngImage(XLImage):
        def __init__(self, raw: bytes):
            self.ref = io.BytesIO(raw)
            self._raw = raw
            self.width, self.height = _png_size(raw)
            self.format = "png"

        def _data(self):
            return self._raw

    return _PngImage(data)

# 签名图在 Excel 中的显示高度（磅→像素按 96dpi 估算），行高随之调整
_SIGN_ROW_HEIGHT = 40
_SIGN_IMG_HEIGHT = 48


def export_issuance(operator: str, where_sql: str = "", params: tuple = ()) -> tuple:
    """导出证件领用记录。签名以图片嵌入对应单元格。

    JOIN personnel_filing 以排除孤儿行（延续既有数据完整性口径）。
    """
    from utils.helpers import get_dict_value

    db = get_db()
    rows = db.execute(
        "SELECT i.*, pf.work_unit FROM cert_issuance i "
        "JOIN personnel_filing pf ON i.personnel_filing_id = pf.id "
        "WHERE 1=1 " + where_sql + " ORDER BY i.issue_date DESC, i.id DESC",
        params,
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "证件领用登记表"
    _style_header(ws, "因私出国（境）证件领用登记表", HEADERS_ISS)

    # 图片对象需在写盘前保持引用，否则可能被 GC 回收
    _keep = []
    for i, row in enumerate(rows, 3):
        types = "、".join(
            get_dict_value("cert_type", c) or c
            for c in (row["cert_types"] or "").split(",") if c
        )
        values = [
            row["work_unit"], row["holder_name"], row["id_number"] or "", types,
            row["cert_nos"] or "", row["issue_date"], row["issuer"], "",
            row["return_date"] or "", row["return_operator"] or "", "",
            _STATUS_LABEL.get(row["status"], row["status"]), row["remarks"] or "",
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=i, column=col, value=val)

        # 第 8 列＝领用签名，第 11 列＝归还签名
        has_img = False
        for col, blob in ((8, row["sign_image"]), (11, row["return_sign_image"])):
            if not blob:
                continue
            try:
                img = _make_png_image(bytes(blob))
                ratio = _SIGN_IMG_HEIGHT / img.height if img.height else 1
                img.height = _SIGN_IMG_HEIGHT
                img.width = max(int(img.width * ratio), 1)
                img.anchor = f"{get_column_letter(col)}{i}"
                ws.add_image(img)
                _keep.append(img)
                has_img = True
            except Exception:
                # 单张签名渲染失败不应中断整表导出
                ws.cell(row=i, column=col, value="[签名图无法读取]")
        if has_img:
            ws.row_dimensions[i].height = _SIGN_ROW_HEIGHT

    _style_data(ws, 3, len(rows) + 2, len(HEADERS_ISS))
    _auto_width(ws, len(HEADERS_ISS))
    # 签名列按图片宽度留白（自动列宽只看文本，会把空单元格压窄）
    for col in (8, 11):
        ws.column_dimensions[get_column_letter(col)].width = 22

    return _save_and_return(ws, "证件领用登记表", operator, [
        "1. 签名为领用/归还时现场手写采集，保存后不可修改；登记有误须作废后重新登记。",
        "2. 证件号码为领用当时的快照，后续修改证照信息不影响本次领用凭证。",
        "3. 状态：已领用（未归还）/ 已归还 / 已作废。",
    ])


# =========================================================================
# 7. 操作日志年度归档
# =========================================================================
HEADERS_LOGS = ["时间（本地）", "操作人", "动作", "对象类型", "对象ID", "详情", "IP", "变更快照(JSON)"]


def export_logs(operator: str, year: str) -> tuple:
    """按年份归档导出操作日志（时间按本地时区换算与过滤）。"""
    from utils.helpers import to_local_time, tz_modifier
    db = get_db()
    tz = tz_modifier()
    rows = db.execute(
        "SELECT * FROM operation_logs "
        "WHERE strftime('%Y', datetime(created_at, ?)) = ? ORDER BY created_at",
        (tz, year),
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年操作日志"
    _style_header(ws, f"操作日志归档（{year} 年）", HEADERS_LOGS)

    for i, row in enumerate(rows, 3):
        values = [
            to_local_time(row["created_at"]), row["operator"], row["action"],
            row["target_type"], row["target_id"], row["detail"] or "",
            row["ip_address"] or "",
            (row["snapshot"] or "") if "snapshot" in row.keys() else "",
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=i, column=col, value=val)

    _style_data(ws, 3, len(rows) + 2, len(HEADERS_LOGS))
    _auto_width(ws, len(HEADERS_LOGS))
    return _save_and_return(ws, f"操作日志归档_{year}年", operator, [
        "1. 时间已按系统配置时区换算为本地时间。",
        "2. 本文件为审计归档副本；数据库中的日志不可删除，仍完整保留。",
    ])
